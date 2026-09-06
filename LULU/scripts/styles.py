"""Shared styling helpers enforcing the GIS color convention:

    BLUE  font -> hard-coded numbers reported by the company (10-K / release)
    BLACK font -> calculations / formulas
    RED   font -> analyst assumptions / inputs

Uses openpyxl. Garamond is applied where available (matches the GIS deck).
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.hyperlink import Hyperlink

FONT_NAME = "Garamond"

BLUE = "0000CC"    # reported
BLACK = "000000"   # formula
RED = "C00000"     # assumption
WHITE = "FFFFFF"
NAVY = "1F2A44"
GREEN = "006100"

# GIS-style palette
DARK = "1F2A44"      # header navy
ACCENT = "C8102E"    # USC-ish cardinal accent
LIGHT = "D9E1F2"     # light blue band
GREY = "F2F2F2"

def font(color=BLACK, bold=False, size=10, italic=False, name=FONT_NAME, underline=None):
    return Font(name=name, color=color, bold=bold, size=size, italic=italic, underline=underline)

def fill(color):
    return PatternFill("solid", fgColor=color)

thin = Side(style="thin", color="BFBFBF")
medium = Side(style="medium", color="000000")
double = Side(style="double", color="000000")

def border(top=None, bottom=None, left=None, right=None):
    return Border(top=top, bottom=bottom, left=left, right=right)

top_border = border(top=thin)
bottom_border = border(bottom=thin)
top_bottom = border(top=thin, bottom=thin)
top_double = border(top=Side(style="thin", color="000000"),
                    bottom=Side(style="double", color="000000"))

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")
left_indent = Alignment(horizontal="left", vertical="center", indent=1)
left_indent2 = Alignment(horizontal="left", vertical="center", indent=2)

# number formats
NUM = '#,##0;(#,##0)'
NUM1 = '#,##0.0;(#,##0.0)'
PCT = '0.0%'
PCT0 = '0%'
MONEY = '$#,##0.00'
MULT = '0.0x'
EPSFMT = '$#,##0.00'

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def write(ws, cell, value, color=BLACK, bold=False, size=10, numfmt=None,
          align=None, fillc=None, bdr=None, italic=False, name=FONT_NAME):
    c = ws[cell]
    c.value = value
    c.font = font(color=color, bold=bold, size=size, italic=italic, name=name)
    if numfmt:
        c.number_format = numfmt
    if align:
        c.alignment = align
    if fillc:
        c.fill = fill(fillc)
    if bdr:
        c.border = bdr
    return c


def format_source_text(label, hint=None):
    """Combine clickable label with exact Ctrl+F search strings for the reader."""
    if hint:
        return f"{label}\nCtrl+F: {hint}"
    return label


def write_link(ws, cell, text, url, color=BLUE, bold=False, size=10, numfmt=None,
               align=None, fillc=None, bdr=None, italic=False, underline="single", hint=None):
    display = format_source_text(text, hint)
    c = write(ws, cell, display, color=color, bold=bold, size=size, numfmt=numfmt,
              align=align or left, fillc=fillc, bdr=bdr, italic=italic)
    c.hyperlink = url
    c.font = font(color=color, bold=bold, size=size, italic=italic, underline=underline)
    if hint:
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        lines = hint.count('\n') + 2
        ws.row_dimensions[c.row].height = max(ws.row_dimensions[c.row].height or 15, min(72, 18 * lines))
    return c


def write_reported(ws, cell, value, source_url=None, bold=False, size=10, numfmt=None,
                   align=None, bdr=None):
    """Blue-font reported figure; hyperlinks to source when URL provided."""
    c = write(ws, cell, value, BLUE, bold=bold, size=size, numfmt=numfmt,
              align=align or right, bdr=bdr)
    if source_url:
        c.hyperlink = source_url
        c.font = font(color=BLUE, bold=bold, size=size, underline="single")
    return c


def write_internal_link(ws, cell, text, location, color=BLUE, size=8, italic=True, hint=None):
    """Clickable link to another cell/tab in this workbook (location e.g. 'WACC!C12')."""
    display = format_source_text(text, hint)
    c = write(ws, cell, display, color, italic=italic, size=size, align=left)
    c.hyperlink = Hyperlink(ref=c.coordinate, location=location)
    c.font = font(color=color, italic=italic, size=size, underline="single")
    if hint:
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        lines = hint.count('\n') + 2
        ws.row_dimensions[c.row].height = max(ws.row_dimensions[c.row].height or 15, min(72, 18 * lines))
    return c


def _hint_row_height(ws, row, hint):
    if hint:
        lines = hint.count('\n') + 2
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, min(72, 18 * lines))


def write_ctrl_f(ws, cell, hint):
    """Dedicated Ctrl+F column — exact strings to locate the number in the source."""
    if not hint:
        return
    c = write(ws, cell, hint, BLACK, italic=True, size=8, align=left)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    _hint_row_height(ws, c.row, hint)


def write_source_with_ctrl_f(ws, source_cell, ctrl_f_cell, label, url, hint=None, size=8):
    """Reported-figure source link + dedicated Ctrl+F column."""
    write_link(ws, source_cell, label, url, color=BLUE, size=size, italic=True)
    write_ctrl_f(ws, ctrl_f_cell, hint)


def write_assumption_docs(ws, row, justify_col, source_col, ctrl_f_col, key, justify_dict, src_dict,
                          internal_location=None, extra_source_col=None, extra_ctrl_f_col=None,
                          extra_label=None, extra_url=None, extra_hint=None, hints=None):
    """Write justification, clickable source, and Ctrl+F proof column."""
    if key in justify_dict:
        c = ws[f"{justify_col}{row}"]
        c.value = justify_dict[key]
        c.font = font(color=BLACK, italic=True, size=8)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hint = (hints or {}).get(key)
    src = src_dict.get(key)
    if internal_location:
        label = src[0] if src else "Model cross-reference"
        write_internal_link(ws, f"{source_col}{row}", label, internal_location, size=9)
    elif src:
        label, url = src
        if url:
            write_link(ws, f"{source_col}{row}", label, url, color=BLUE, size=9,
                       italic=True, underline="single")
        else:
            write(ws, f"{source_col}{row}", label, BLACK, italic=True, size=8, align=left)
    write_ctrl_f(ws, f"{ctrl_f_col}{row}", hint)
    if extra_source_col and extra_label and extra_url:
        write_link(ws, f"{extra_source_col}{row}", extra_label, extra_url,
                   color=BLUE, size=8, italic=True, underline="single")
        if extra_ctrl_f_col and extra_hint:
            write_ctrl_f(ws, f"{extra_ctrl_f_col}{row}", extra_hint)


def append_assumption_docs(ws, row, justify_col, source_col, ctrl_f_col, key, justify_dict, src_dict,
                           hints=None, prefix="Also"):
    """Add a second assumption block (e.g. repurchase price) below the first in B/C/D."""
    if key in justify_dict:
        c = ws[f"{justify_col}{row}"]
        extra = f"{prefix}: {justify_dict[key]}"
        c.value = f"{c.value}\n{extra}" if c.value else extra
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    hint = (hints or {}).get(key)
    if hint:
        c = ws[f"{ctrl_f_col}{row}"]
        extra = f"{prefix}: {hint}"
        c.value = f"{c.value}\n{extra}" if c.value else extra
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        _hint_row_height(ws, row, c.value)
    src = src_dict.get(key)
    if src:
        label, url = src
        c = ws[f"{source_col}{row}"]
        extra = f"{prefix}: {label}"
        c.value = f"{c.value}\n{extra}" if c.value else extra
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


def write_ff_dual_docs(ws, row, lo_key, hi_key, justify_col, source_col, ctrl_f_col,
                       justify_dict, src_dict, hints=None):
    """Football-field row: lo + hi assumption docs combined in visible B/C/D columns."""
    lo_j = justify_dict.get(lo_key, "")
    hi_j = justify_dict.get(hi_key, "")
    write(ws, f"{justify_col}{row}", f"Lo: {lo_j}\nHi: {hi_j}", BLACK, italic=True, size=8, align=left)
    ws[f"{justify_col}{row}"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    lo_src = src_dict.get(lo_key)
    hi_src = src_dict.get(hi_key)
    if lo_src and lo_src[1]:
        write_link(ws, f"{source_col}{row}", f"Lo: {lo_src[0]}", lo_src[1], color=BLUE, size=8, italic=True)
        if hi_src:
            c = ws[f"{source_col}{row}"]
            c.value = f"{c.value}\nHi: {hi_src[0]}"
            c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    lo_hint = (hints or {}).get(lo_key, "")
    hi_hint = (hints or {}).get(hi_key, "")
    write_ctrl_f(ws, f"{ctrl_f_col}{row}", f"Lo: {lo_hint}\nHi: {hi_hint}" if hi_hint else lo_hint)
