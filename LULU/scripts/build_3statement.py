"""Builds LULU_3_Statement_Model.xlsx from scratch (no third-party template).

Historicals (FY2022-FY2025) are hard-coded reported figures (BLUE).
Projections (FY2026E-FY2030E) are live Excel formulas (BLACK) driven by
analyst assumptions on the Assumptions tab (RED). The statements are linked;
the balance sheet ties by construction (cash is the plug from the CFS).
Units: US$ thousands unless noted.
"""
import os
from openpyxl import Workbook
import styles as S
from styles import write, write_link, write_reported, write_assumption_docs, write_ctrl_f, append_assumption_docs, NUM, PCT, MONEY, EPSFMT
import data as D

OUT = os.path.join(os.path.dirname(__file__), "..", "LULU_3_Statement_Model.xlsx")

HIST = D.HIST_YEARS
PROJ = D.PROJ_YEARS
YEARS = HIST + PROJ
DJ, DS, DC = "B", "C", "D"
HCOLS = ['E', 'F', 'G', 'H']
PCOLS = ['I', 'J', 'K', 'L', 'M']
COLS = HCOLS + PCOLS
COL = dict(zip(YEARS, COLS))
YEAR_OF = {c: y for y, c in COL.items()}
PREV = {c: COLS[i-1] for i, c in enumerate(COLS) if i > 0}

A, ISN, BSN, CFN = "Assumptions", "Income Statement", "Balance Sheet", "Cash Flow"
WIDTHS = {'A': 32, 'B': 24, 'C': 16, 'D': 36, 'E': 10, 'F': 10, 'G': 10, 'H': 10,
          'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10}

wb = Workbook()


def year_header(ws, title):
    write(ws, 'A1', title, S.WHITE, bold=True, size=12, fillc=S.DARK, align=S.left_indent)
    for y in YEARS:
        write(ws, f'{COL[y]}1', y, S.WHITE, bold=True, size=10, align=S.center,
              fillc=(S.ACCENT if y in PROJ else S.DARK))
    ws.row_dimensions[1].height = 16
    write(ws, f'{DJ}2', "Justification (~20 words)", S.ACCENT, bold=True, size=9, align=S.left_indent)
    write(ws, f'{DS}2', "Source (click)", S.ACCENT, bold=True, size=9, align=S.left_indent)
    write(ws, f'{DC}2', "Ctrl+F (prove number)", S.ACCENT, bold=True, size=9, align=S.left_indent)
    for y in HIST:
        write_link(ws, f'{COL[y]}2', "10-K", D.filing_url(y), color=S.BLUE, size=8, align=S.center)
    write(ws, f'{PCOLS[0]}2', "Projected", S.ACCENT, italic=True, size=8, align=S.center)
    ws.freeze_panes = 'E3'


def write_row_docs(ws, row, justify_key, extra_key=None):
    if justify_key:
        write_assumption_docs(ws, row, DJ, DS, DC, justify_key, D.JUST, D.ASSUMPTION_SRC,
                              hints=D.SOURCE_HINT)
        if extra_key:
            append_assumption_docs(ws, row, DJ, DS, DC, extra_key, D.JUST, D.ASSUMPTION_SRC,
                                   hints=D.SOURCE_HINT)


# ---------------------------------------------------------------- COVER
cov = wb.active
cov.title = "Cover"
cov.sheet_view.showGridLines = False
S.set_col_widths(cov, {'A': 3, 'B': 100})
write(cov, 'B2', "GLOBAL INVESTMENT SOCIETY  |  INVESTMENT RESEARCH DIVISION", S.WHITE, bold=True, size=13, fillc=S.DARK)
cov['B2'].alignment = S.Alignment(horizontal='left', vertical='center', indent=1)
cov.row_dimensions[2].height = 26
write(cov, 'B4', "lululemon athletica inc. (NASDAQ: LULU)", S.DARK, bold=True, size=20)
write(cov, 'B5', "Integrated Three-Statement Operating Model", S.ACCENT, bold=True, size=13)
write(cov, 'B7', "Historical figures (FY2022\u2013FY2025) per company SEC filings (Form 10-K).", S.BLACK, size=11)
write(cov, 'B8', "Fiscal year ends late January / early February; FY2025 ended February 1, 2026.", S.BLACK, size=11)
write(cov, 'B10', "FONT / COLOR CONVENTION", S.DARK, bold=True, size=12)
write(cov, 'B11', "Blue font  =  figures reported by the company (click value or 10-K link for source)", S.BLUE, bold=True, size=11)
write(cov, 'B12', "Black font  =  calculations / formulas", S.BLACK, bold=True, size=11)
write(cov, 'B13', "Red font  =  analyst assumptions — cols B/C/D on every tab: Justification | Source | Ctrl+F", S.RED, bold=True, size=11)
write(cov, 'B15', "SOURCES", S.DARK, bold=True, size=12)
write_link(cov, 'B16', "SEC EDGAR filings, CIK 0001397187 (Forms 10-K)", D.SOURCES["edgar_xbrl"],
           color=S.BLUE, size=10, hint=D.COVER_HINTS["edgar_xbrl"])
write_link(cov, 'B17', "FY2025 Form 10-K (ended Feb 1, 2026)", D.filing_url("FY2025"),
           color=S.BLUE, size=10, hint=D.COVER_HINTS["filing_fy2025"])
write_link(cov, 'B18', "Q2 FY2026 results & FY2026 guidance (Sep 3, 2026 earnings release)", D.SOURCES["earnings_sep2026"],
           color=S.BLUE, size=10, hint=D.COVER_HINTS["earnings_sep2026"])
write(cov, 'B19', "TABS", S.DARK, bold=True, size=12)
write(cov, 'B20', "Assumptions  \u2022  Income Statement  \u2022  Balance Sheet  \u2022  Cash Flow", S.BLACK, size=10)
write(cov, 'B22', "Built from scratch for the GIS IR selection assignment. Units: US$ thousands unless noted.", S.BLACK, italic=True, size=9)

# ---------------------------------------------------------------- ASSUMPTIONS
asum = wb.create_sheet(A)
asum.sheet_view.showGridLines = False
S.set_col_widths(asum, WIDTHS)
year_header(asum, "OPERATING ASSUMPTIONS & DRIVERS")
for y in YEARS:
    write(asum, f'{COL[y]}2', D.HIST_END.get(y, "proj."), S.BLACK, italic=True, size=7, align=S.center)

AR = {}
r = [4]

def a_section(title):
    write(asum, f'A{r[0]}', title, S.DARK, bold=True, size=10, fillc=S.GREY)
    for c in COLS:
        asum[f'{c}{r[0]}'].fill = S.fill(S.GREY)
    r[0] += 1

def a_row(key, label, hist_vals, proj_vals, fmt=PCT, justify_key=""):
    AR[key] = r[0]
    write(asum, f'A{r[0]}', label, S.BLACK, size=10, align=S.left_indent)
    for i, y in enumerate(HIST):
        if hist_vals[i] is not None:
            write(asum, f'{COL[y]}{r[0]}', hist_vals[i], S.BLACK, size=10, numfmt=fmt, align=S.right)
    for i, y in enumerate(PROJ):
        write(asum, f'{COL[y]}{r[0]}', proj_vals[i], S.RED, size=10, numfmt=fmt, align=S.right)
    if justify_key:
        write_row_docs(asum, r[0], justify_key)
    r[0] += 1

rev, cogs = D.IS['revenue'], D.IS['cogs']
asum.freeze_panes = 'E4'
a_section("GROWTH & MARGINS")
a_row('rev_growth', "Revenue growth %",
      [None, rev['FY2023']/rev['FY2022']-1, rev['FY2024']/rev['FY2023']-1, rev['FY2025']/rev['FY2024']-1],
      [-0.061, 0.026, 0.023, 0.023, 0.023], justify_key="3s_rev_growth")
a_row('gm', "Gross margin %", [D.IS['gross_profit'][y]/rev[y] for y in HIST],
      [0.565, 0.570, 0.575, 0.575, 0.580], justify_key="3s_gm")
a_row('sga_pct', "SG&A % of revenue", [D.IS['sga'][y]/rev[y] for y in HIST],
      [0.425, 0.415, 0.405, 0.400, 0.395], justify_key="3s_sga_pct")
a_row('other_opex', "Amortization / other opex ($)", [D.IS['other_opex'][y] for y in HIST], [7000]*5, fmt=NUM,
      justify_key="3s_other_opex")
a_row('other_inc', "Other income, net ($)", [D.IS['other_income'][y] for y in HIST],
      [45000, 40000, 35000, 30000, 25000], fmt=NUM, justify_key="3s_other_inc")
a_row('tax_rate', "Effective tax rate %", [D.IS['tax'][y]/D.IS['pretax_income'][y] for y in HIST], [0.300]*5,
      justify_key="3s_tax_rate")

a_section("CAPITAL & NON-CASH ITEMS")
a_row('da_pct', "D&A % of revenue", [D.CF['d_and_a'][y]/rev[y] for y in HIST], [0.046, 0.046, 0.045, 0.045, 0.045],
      justify_key="3s_da_pct")
a_row('capex_pct', "Capex % of revenue", [D.CF['capex'][y]/rev[y] for y in HIST], [0.070, 0.060, 0.055, 0.050, 0.050],
      justify_key="3s_capex_pct")
a_row('sbc', "Stock-based compensation ($)", [D.CF['sbc'][y] for y in HIST], [62000]*5, fmt=NUM,
      justify_key="3s_sbc")

a_section("WORKING CAPITAL & BALANCE SHEET DRIVERS")
oca = {y: D.BS['current_assets'][y]-D.BS['cash'][y]-D.BS['inventories'][y] for y in HIST}
onca = {y: D.BS['total_assets'][y]-D.BS['current_assets'][y]-D.BS['ppe_net'][y]-D.BS['rou_asset'][y]-D.BS['goodwill_intang'][y] for y in HIST}
ocl = {y: D.BS['current_liab'][y]-D.BS['accounts_payable'][y]-D.BS['accrued_liab'][y]-D.BS['op_lease_cur'][y] for y in HIST}
oncl = {y: D.BS['total_liab'][y]-D.BS['current_liab'][y]-D.BS['op_lease_noncur'][y]-D.BS['deferred_tax'][y] for y in HIST}
a_row('inv_pct', "Inventories % of COGS", [D.BS['inventories'][y]/cogs[y] for y in HIST], [0.340, 0.335, 0.330, 0.330, 0.330],
      justify_key="3s_inv_pct")
a_row('ap_pct', "Accounts payable % of COGS", [D.BS['accounts_payable'][y]/cogs[y] for y in HIST], [0.068]*5,
      justify_key="3s_ap_pct")
a_row('accr_pct', "Accrued liabilities % of revenue", [D.BS['accrued_liab'][y]/rev[y] for y in HIST], [0.058]*5,
      justify_key="3s_accr_pct")
a_row('oca_pct', "Other current assets % of revenue", [oca[y]/rev[y] for y in HIST], [0.068]*5,
      justify_key="3s_oca_pct")
a_row('rou_pct', "Operating lease ROU asset % of revenue", [D.BS['rou_asset'][y]/rev[y] for y in HIST], [0.147]*5,
      justify_key="3s_rou_pct")
a_row('onca_pct', "Other non-current assets % of revenue", [onca[y]/rev[y] for y in HIST], [0.030]*5,
      justify_key="3s_onca_pct")
a_row('olc_pct', "Op lease liab (current) % of revenue", [D.BS['op_lease_cur'][y]/rev[y] for y in HIST], [0.027]*5,
      justify_key="3s_olc_pct")
a_row('olnc_pct', "Op lease liab (non-current) % of revenue", [D.BS['op_lease_noncur'][y]/rev[y] for y in HIST], [0.135]*5,
      justify_key="3s_olnc_pct")
a_row('ocl_pct', "Other current liabilities % of revenue", [ocl[y]/rev[y] for y in HIST], [0.055]*5,
      justify_key="3s_ocl_pct")
a_row('oncl_pct', "Other non-current liab % of revenue", [oncl[y]/rev[y] for y in HIST], [0.005]*5,
      justify_key="3s_oncl_pct")

a_section("CAPITAL RETURN")
a_row('buyback', "Share repurchases ($)", [D.CF['buybacks'][y] for y in HIST], [500000]*5, fmt=NUM,
      justify_key="3s_buyback")
a_row('rep_price', "Avg repurchase price ($/sh)", [None]*4, [100, 108, 115, 122, 130], fmt=MONEY,
      justify_key="3s_rep_price")

write(asum, f'A{r[0]+1}', "Historical columns = derived ratios (black); projection columns = analyst inputs (red).",
      S.BLACK, italic=True, size=8, align=S.left_indent)


def ar(key, col):
    return f"'{A}'!{col}{AR[key]}"

# ---------------------------------------------------------------- INCOME STATEMENT
is_ = wb.create_sheet(ISN)
is_.sheet_view.showGridLines = False
S.set_col_widths(is_, WIDTHS)
year_header(is_, "INCOME STATEMENT  (US$ thousands)")
IR = {}
def isref(key, col):
    return f"'{ISN}'!{col}{IR[key]}"

def is_reported(key, label, hist_dict, proj_fn, bold=False, top=False, dbl=False, fmt=NUM, justify_key="", extra_justify_key=""):
    IR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(is_, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    write_row_docs(is_, r[0], justify_key, extra_justify_key)
    for y in HIST:
        write_reported(is_, f'{COL[y]}{r[0]}', hist_dict[y], D.filing_url(y),
                       bold=bold, size=10, numfmt=fmt, bdr=bdr)
    for y in PROJ:
        c = COL[y]
        write(is_, f'{c}{r[0]}', proj_fn(c), S.RED if justify_key else S.BLACK,
              bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    r[0] += 1

def is_calc(key, label, fn, bold=False, top=False, dbl=False, fmt=NUM):
    IR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(is_, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    for y in YEARS:
        c = COL[y]
        write(is_, f'{c}{r[0]}', fn(c), S.BLACK, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    r[0] += 1

r = [4]
is_reported('rev', "Net revenue", D.IS['revenue'], lambda c: f"={PREV[c]}{IR['rev']}*(1+{ar('rev_growth', c)})",
            justify_key="3s_rev_growth")
is_reported('cogs', "Cost of goods sold", D.IS['cogs'], lambda c: f"={c}{IR['rev']}*(1-{ar('gm', c)})",
            justify_key="3s_gm")
is_calc('gp', "Gross profit", lambda c: f"={c}{IR['rev']}-{c}{IR['cogs']}", bold=True, top=True)
is_reported('sga', "Selling, general & administrative", D.IS['sga'], lambda c: f"={c}{IR['rev']}*{ar('sga_pct', c)}",
            justify_key="3s_sga_pct")
is_reported('oopex', "Amortization of intangibles / other", D.IS['other_opex'], lambda c: f"={ar('other_opex', c)}",
            fmt=NUM, justify_key="3s_other_opex")
is_calc('ebit', "Operating income (EBIT)", lambda c: f"={c}{IR['gp']}-{c}{IR['sga']}-{c}{IR['oopex']}", bold=True, top=True)
is_reported('oinc', "Other income, net", D.IS['other_income'], lambda c: f"={ar('other_inc', c)}",
            fmt=NUM, justify_key="3s_other_inc")
is_calc('pretax', "Pre-tax income", lambda c: f"={c}{IR['ebit']}+{c}{IR['oinc']}", bold=True, top=True)
is_reported('tax', "Income tax expense", D.IS['tax'], lambda c: f"={c}{IR['pretax']}*{ar('tax_rate', c)}",
            justify_key="3s_tax_rate")
is_calc('ni', "Net income", lambda c: f"={c}{IR['pretax']}-{c}{IR['tax']}", bold=True, top=True, dbl=True)
r[0] += 1
is_reported('sh', "Diluted weighted-avg shares (000)", D.IS['diluted_shares'],
            lambda c: f"={PREV[c]}{IR['sh']}-{ar('buyback', c)}/{ar('rep_price', c)}",
            justify_key="3s_buyback", extra_justify_key="3s_rep_price")
IR['eps'] = r[0]
write(is_, f'A{r[0]}', "Diluted EPS ($)", S.DARK, bold=True, size=10, align=S.left_indent)
for y in HIST:
    write_reported(is_, f'{COL[y]}{r[0]}', D.IS['diluted_eps'][y], D.filing_url(y),
                   bold=True, size=10, numfmt=EPSFMT)
for y in PROJ:
    c = COL[y]
    write(is_, f'{c}{r[0]}', f"={c}{IR['ni']}/{c}{IR['sh']}", S.BLACK, bold=True, size=10, numfmt=EPSFMT, align=S.right)
r[0] += 2
write(is_, f'A{r[0]}', "Margins & growth", S.DARK, bold=True, size=9, fillc=S.GREY)
for c in COLS:
    is_[f'{c}{r[0]}'].fill = S.fill(S.GREY)
r[0] += 1
def is_memo(label, fn, growth=False):
    write(is_, f'A{r[0]}', label, S.BLACK, italic=True, size=9, align=S.left_indent)
    for i, y in enumerate(YEARS):
        c = COL[y]
        if growth and i == 0:
            write(is_, f'{c}{r[0]}', "n/a", S.BLACK, italic=True, size=9, align=S.right)
        else:
            write(is_, f'{c}{r[0]}', fn(c), S.BLACK, italic=True, size=9, numfmt=PCT, align=S.right)
    r[0] += 1
is_memo("Gross margin %", lambda c: f"={c}{IR['gp']}/{c}{IR['rev']}")
is_memo("Operating margin %", lambda c: f"={c}{IR['ebit']}/{c}{IR['rev']}")
is_memo("Net margin %", lambda c: f"={c}{IR['ni']}/{c}{IR['rev']}")
is_memo("Revenue growth %", lambda c: f"={c}{IR['rev']}/{PREV[c]}{IR['rev']}-1", growth=True)

# ---------------------------------------------------------------- BALANCE SHEET
bs = wb.create_sheet(BSN)
bs.sheet_view.showGridLines = False
S.set_col_widths(bs, WIDTHS)
year_header(bs, "BALANCE SHEET  (US$ thousands)")
BR = {}

def bs_band(text):
    write(bs, f'A{r[0]}', text, S.DARK, bold=True, size=10, fillc=S.GREY)
    for c in COLS:
        bs[f'{c}{r[0]}'].fill = S.fill(S.GREY)
    r[0] += 1

def bs_sub(text):
    write(bs, f'A{r[0]}', text, S.BLACK, bold=True, italic=True, size=10, align=S.left_indent)
    r[0] += 1

def bs_rep(key, label, hist_dict, proj_fn, bold=False, top=False, dbl=False, justify_key=""):
    BR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(bs, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    write_row_docs(bs, r[0], justify_key)
    for y in HIST:
        write_reported(bs, f'{COL[y]}{r[0]}', hist_dict[y], D.filing_url(y),
                       bold=bold, size=10, numfmt=NUM, bdr=bdr)
    for y in PROJ:
        c = COL[y]
        col = S.RED if justify_key else S.BLACK
        write(bs, f'{c}{r[0]}', proj_fn(c), col, bold=bold, size=10, numfmt=NUM, align=S.right, bdr=bdr)
    r[0] += 1

def bs_plug(key, label, hist_valdict, proj_fn, justify_key=""):
    """hist = black computed VALUE; proj = black/red formula."""
    BR[key] = r[0]
    write(bs, f'A{r[0]}', label, S.BLACK, size=10, align=S.left_indent)
    write_row_docs(bs, r[0], justify_key)
    for y in HIST:
        write(bs, f'{COL[y]}{r[0]}', hist_valdict[y], S.BLACK, size=10, numfmt=NUM, align=S.right)
    for y in PROJ:
        c = COL[y]
        col = S.RED if justify_key else S.BLACK
        write(bs, f'{c}{r[0]}', proj_fn(c), col, size=10, numfmt=NUM, align=S.right)
    r[0] += 1

def bs_totcalc(key, label, fn, bold=True, top=True, dbl=False):
    BR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(bs, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    for y in YEARS:
        c = COL[y]
        write(bs, f'{c}{r[0]}', fn(c), S.BLACK, bold=bold, size=10, numfmt=NUM, align=S.right, bdr=bdr)
    r[0] += 1

def isr(key, col):
    return f"'{ISN}'!{col}{IR[key]}"

r = [4]
bs_band("ASSETS")
bs_sub("Current assets:")
BR['cash'] = r[0]
write(bs, f'A{r[0]}', "Cash & cash equivalents", S.BLACK, size=10, align=S.left_indent)
for y in HIST:
    write_reported(bs, f'{COL[y]}{r[0]}', D.BS['cash'][y], D.filing_url(y), size=10, numfmt=NUM)
r[0] += 1  # projected cash filled after CF is built
bs_rep('inv', "Inventories", D.BS['inventories'], lambda c: f"={isr('cogs', c)}*{ar('inv_pct', c)}", justify_key="3s_inv_pct")
bs_plug('oca', "Other current assets", oca, lambda c: f"={isr('rev', c)}*{ar('oca_pct', c)}", justify_key="3s_oca_pct")
bs_rep('tca', "Total current assets", D.BS['current_assets'],
       lambda c: f"={c}{BR['cash']}+{c}{BR['inv']}+{c}{BR['oca']}", bold=True, top=True)
bs_sub("Non-current assets:")
bs_rep('ppe', "Property & equipment, net", D.BS['ppe_net'],
       lambda c: f"={PREV[c]}{BR['ppe']}+{isr('rev', c)}*{ar('capex_pct', c)}-{isr('rev', c)}*{ar('da_pct', c)}",
       justify_key="3s_capex_pct")
bs_rep('rou', "Operating lease right-of-use assets", D.BS['rou_asset'], lambda c: f"={isr('rev', c)}*{ar('rou_pct', c)}", justify_key="3s_rou_pct")
bs_rep('gwi', "Goodwill & intangible assets", D.BS['goodwill_intang'], lambda c: f"={PREV[c]}{BR['gwi']}")
bs_plug('onca', "Other non-current assets", onca, lambda c: f"={isr('rev', c)}*{ar('onca_pct', c)}", justify_key="3s_onca_pct")
bs_rep('ta', "TOTAL ASSETS", D.BS['total_assets'],
       lambda c: f"={c}{BR['tca']}+{c}{BR['ppe']}+{c}{BR['rou']}+{c}{BR['gwi']}+{c}{BR['onca']}", bold=True, top=True, dbl=True)
r[0] += 1
bs_band("LIABILITIES & EQUITY")
bs_sub("Current liabilities:")
bs_rep('ap', "Accounts payable", D.BS['accounts_payable'], lambda c: f"={isr('cogs', c)}*{ar('ap_pct', c)}", justify_key="3s_ap_pct")
bs_rep('accr', "Accrued liabilities", D.BS['accrued_liab'], lambda c: f"={isr('rev', c)}*{ar('accr_pct', c)}", justify_key="3s_accr_pct")
bs_rep('olc', "Operating lease liabilities (current)", D.BS['op_lease_cur'], lambda c: f"={isr('rev', c)}*{ar('olc_pct', c)}", justify_key="3s_olc_pct")
bs_plug('ocl', "Other current liabilities", ocl, lambda c: f"={isr('rev', c)}*{ar('ocl_pct', c)}", justify_key="3s_ocl_pct")
bs_rep('tcl', "Total current liabilities", D.BS['current_liab'],
       lambda c: f"={c}{BR['ap']}+{c}{BR['accr']}+{c}{BR['olc']}+{c}{BR['ocl']}", bold=True, top=True)
bs_sub("Non-current liabilities:")
bs_rep('olnc', "Operating lease liabilities (non-current)", D.BS['op_lease_noncur'], lambda c: f"={isr('rev', c)}*{ar('olnc_pct', c)}", justify_key="3s_olnc_pct")
bs_rep('dtl', "Deferred income taxes", D.BS['deferred_tax'], lambda c: f"={PREV[c]}{BR['dtl']}")
bs_plug('oncl', "Other non-current liabilities", oncl, lambda c: f"={isr('rev', c)}*{ar('oncl_pct', c)}", justify_key="3s_oncl_pct")
bs_rep('tl', "TOTAL LIABILITIES", D.BS['total_liab'],
       lambda c: f"={c}{BR['tcl']}+{c}{BR['olnc']}+{c}{BR['dtl']}+{c}{BR['oncl']}", bold=True, top=True)
r[0] += 1
bs_sub("Shareholders' equity:")
bs_rep('capic', "Common stock & additional paid-in capital", D.BS['common_apic'],
       lambda c: f"={PREV[c]}{BR['capic']}+{ar('sbc', c)}", justify_key="3s_sbc")
bs_rep('re', "Retained earnings", D.BS['retained_earn'],
       lambda c: f"={PREV[c]}{BR['re']}+{isr('ni', c)}-{ar('buyback', c)}", justify_key="3s_buyback")
bs_rep('aoci', "Accumulated other comprehensive loss", D.BS['aoci'], lambda c: f"={PREV[c]}{BR['aoci']}")
bs_rep('te', "TOTAL SHAREHOLDERS' EQUITY", D.BS['total_equity'],
       lambda c: f"={c}{BR['capic']}+{c}{BR['re']}+{c}{BR['aoci']}", bold=True, top=True)
bs_totcalc('tle', "TOTAL LIABILITIES & EQUITY", lambda c: f"={c}{BR['tl']}+{c}{BR['te']}", dbl=True)
r[0] += 1
write(bs, f'A{r[0]}', "Balance check (Assets \u2212 L&E)", S.BLACK, italic=True, size=9, align=S.left_indent)
for y in YEARS:
    c = COL[y]
    write(bs, f'{c}{r[0]}', f"={c}{BR['ta']}-{c}{BR['tle']}", S.BLACK, italic=True, size=9, numfmt=NUM, align=S.right)
r[0] += 1

# ---------------------------------------------------------------- CASH FLOW
cf = wb.create_sheet(CFN)
cf.sheet_view.showGridLines = False
S.set_col_widths(cf, WIDTHS)
year_header(cf, "CASH FLOW STATEMENT  (US$ thousands)")
CR = {}
def bref(key, col):
    return f"'{BSN}'!{col}{BR[key]}"
def delta(bkey, c):
    return f"({bref(bkey, c)}-{bref(bkey, PREV[c])})"
def isr2(key, col):
    return f"'{ISN}'!{col}{IR[key]}"

def cf_band(text):
    write(cf, f'A{r[0]}', text, S.DARK, bold=True, size=10, fillc=S.GREY)
    for c in COLS:
        cf[f'{c}{r[0]}'].fill = S.fill(S.GREY)
    r[0] += 1

def cf_row(key, label, hist_dict, proj_fn, bold=False, top=False, dbl=False, blue=True, justify_key=""):
    CR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(cf, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    write_row_docs(cf, r[0], justify_key)
    for y in HIST:
        if hist_dict is None:
            write(cf, f'{COL[y]}{r[0]}', "\u2014", S.BLACK, size=10, align=S.right, bdr=bdr)
        else:
            write_reported(cf, f'{COL[y]}{r[0]}', hist_dict[y], D.filing_url(y),
                           bold=bold, size=10, numfmt=NUM, bdr=bdr)
    for y in PROJ:
        c = COL[y]
        col = S.RED if justify_key else S.BLACK
        write(cf, f'{c}{r[0]}', proj_fn(c) if proj_fn else "\u2014", col, bold=bold, size=10, numfmt=NUM, align=S.right, bdr=bdr)
    r[0] += 1

r = [4]
cf_band("OPERATING ACTIVITIES")
cf_row('ni', "Net income", D.CF['net_income'], lambda c: f"={isr2('ni', c)}")
cf_row('da', "Depreciation & amortization", D.CF['d_and_a'], lambda c: f"={isr2('rev', c)}*{ar('da_pct', c)}", justify_key="3s_da_pct")
cf_row('sbc', "Stock-based compensation", D.CF['sbc'], lambda c: f"={ar('sbc', c)}", justify_key="3s_sbc")
wc_start = r[0]
cf_row('d_inv', "  (Incr.)/decr. in inventories", None, lambda c: f"=-{delta('inv', c)}", justify_key="3s_inv_pct")
cf_row('d_oca', "  (Incr.)/decr. in other current assets", None, lambda c: f"=-{delta('oca', c)}", justify_key="3s_oca_pct")
cf_row('d_rou', "  (Incr.)/decr. in ROU assets", None, lambda c: f"=-{delta('rou', c)}", justify_key="3s_rou_pct")
cf_row('d_onca', "  (Incr.)/decr. in other non-current assets", None, lambda c: f"=-{delta('onca', c)}", justify_key="3s_onca_pct")
cf_row('d_ap', "  Incr./(decr.) in accounts payable", None, lambda c: f"={delta('ap', c)}", justify_key="3s_ap_pct")
cf_row('d_accr', "  Incr./(decr.) in accrued liabilities", None, lambda c: f"={delta('accr', c)}", justify_key="3s_accr_pct")
cf_row('d_ocl', "  Incr./(decr.) in other current liabilities", None, lambda c: f"={delta('ocl', c)}", justify_key="3s_ocl_pct")
cf_row('d_ol', "  Incr./(decr.) in operating lease liab.", None, lambda c: f"={delta('olc', c)}+{delta('olnc', c)}", justify_key="3s_olc_pct")
cf_row('d_dtl', "  Incr./(decr.) in deferred income taxes", None, lambda c: f"={delta('dtl', c)}")
cf_row('d_oncl', "  Incr./(decr.) in other non-current liab.", None, lambda c: f"={delta('oncl', c)}")
wc_end = r[0] - 1
# historical working-capital reconciling line (black formula referencing reported CFO below)
cfo_row = r[0] + 1  # cf_row for cfo will be next
write(cf, f'A{r[0]}', "  Working capital & other (reported, net)", S.BLACK, size=10, align=S.left_indent)
for y in HIST:
    c = COL[y]
    write(cf, f'{c}{r[0]}', f"={c}{cfo_row}-{c}{CR['ni']}-{c}{CR['da']}-{c}{CR['sbc']}", S.BLACK, size=10, numfmt=NUM, align=S.right)
r[0] += 1
cf_row('cfo', "Net cash from operating activities", D.CF['cfo'],
       lambda c: f"={c}{CR['ni']}+{c}{CR['da']}+{c}{CR['sbc']}+SUM({c}{wc_start}:{c}{wc_end})",
       bold=True, top=True)
assert CR['cfo'] == cfo_row, (CR['cfo'], cfo_row)
r[0] += 1
cf_band("INVESTING ACTIVITIES")
cf_row('capex', "Capital expenditures", {y: -D.CF['capex'][y] for y in HIST},
       lambda c: f"=-{isr2('rev', c)}*{ar('capex_pct', c)}", justify_key="3s_capex_pct")
cf_row('cfi', "Net cash from investing activities", D.CF['cfi'], lambda c: f"={c}{CR['capex']}", bold=True, top=True)
write(cf, f'A{r[0]}', "  (Projections: investing \u2248 capex; excl. M&A / securities)", S.BLACK, italic=True, size=8, align=S.left_indent)
r[0] += 2
cf_band("FINANCING ACTIVITIES")
cf_row('buyback', "Repurchase of common stock", {y: -D.CF['buybacks'][y] for y in HIST},
       lambda c: f"=-{ar('buyback', c)}", justify_key="3s_buyback")
cf_row('cff', "Net cash from financing activities", D.CF['cff'], lambda c: f"={c}{CR['buyback']}", bold=True, top=True)
write(cf, f'A{r[0]}', "  (Projections: financing \u2248 buybacks; SBC non-cash in equity)", S.BLACK, italic=True, size=8, align=S.left_indent)
r[0] += 2
CR['netchg'] = r[0]
write(cf, f'A{r[0]}', "Net change in cash", S.DARK, bold=True, size=10, align=S.left_indent)
for y in YEARS:
    c = COL[y]
    write(cf, f'{c}{r[0]}', f"={c}{CR['cfo']}+{c}{CR['cfi']}+{c}{CR['cff']}", S.BLACK, bold=True, size=10, numfmt=NUM, align=S.right, bdr=S.top_border)
r[0] += 1
begcash_row = r[0]
endcash_row = r[0] + 1
CR['begcash'] = begcash_row
write(cf, f'A{begcash_row}', "Cash, beginning of year", S.BLACK, size=10, align=S.left_indent)
begmap = {'FY2022': 1150517, 'FY2023': D.BS['cash']['FY2022'], 'FY2024': D.BS['cash']['FY2023'], 'FY2025': D.BS['cash']['FY2024']}
for y in HIST:
    write_reported(cf, f'{COL[y]}{begcash_row}', begmap[y], D.filing_url(y), size=10, numfmt=NUM)
write(cf, f"G{begcash_row}", f"='{BSN}'!F{BR['cash']}", S.BLACK, size=10, numfmt=NUM, align=S.right)
for y in PROJ[1:]:
    c = COL[y]
    write(cf, f'{c}{begcash_row}', f"={PREV[c]}{endcash_row}", S.BLACK, size=10, numfmt=NUM, align=S.right)
CR['endcash'] = endcash_row
write(cf, f'A{endcash_row}', "Cash, end of year", S.DARK, bold=True, size=10, align=S.left_indent)
for y in HIST:
    write_reported(cf, f'{COL[y]}{endcash_row}', D.BS['cash'][y], D.filing_url(y),
                   bold=True, size=10, numfmt=NUM, bdr=S.top_double)
for y in PROJ:
    c = COL[y]
    write(cf, f'{c}{endcash_row}', f"={c}{begcash_row}+{c}{CR['netchg']}", S.BLACK, bold=True, size=10, numfmt=NUM, align=S.right, bdr=S.top_double)

# link BS projected cash to CF ending cash
for y in PROJ:
    c = COL[y]
    write(bs, f'{c}{BR["cash"]}', f"='{CFN}'!{c}{endcash_row}", S.BLACK, size=10, numfmt=NUM, align=S.right)

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("Saved", os.path.abspath(OUT))
