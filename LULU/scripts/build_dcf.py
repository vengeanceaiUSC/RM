"""Builds LULU_DCF_Valuation_Model.xlsx from scratch (no third-party template).

Unlevered discounted-cash-flow valuation of lululemon athletica inc.
FY2025 (ended Feb 1, 2026) reported figures anchor the model (BLUE); the
five-year forecast is live formulas (BLACK) driven by analyst assumptions
(RED). Includes a WACC build, Gordon-growth and exit-multiple terminal value,
an EV -> equity bridge, a WACC x terminal-growth sensitivity grid, bull/base/
bear scenarios, and a comps-based football field.
Units: US$ thousands unless noted.
"""
import os
from openpyxl import Workbook
import styles as S
from styles import write, write_link, write_reported, write_assumption_docs, write_internal_link, write_ctrl_f, write_source_with_ctrl_f, write_ff_dual_docs, append_assumption_docs, NUM, PCT, MONEY, MULT, EPSFMT
import data as D

OUT = os.path.join(os.path.dirname(__file__), "..", "LULU_DCF_Valuation_Model.xlsx")
wb = Workbook()

# Documentation columns B/C/D are always visible (no horizontal scroll needed)
DJ, DS, DC = "B", "C", "D"
FY = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
FCOLS = ["F", "G", "H", "I", "J"]     # forecast columns (after FY25 col E)
FCOL = dict(zip(FY, FCOLS))
FY25_COL = "E"
CHK_COL = "K"
SCEN_COLS = ["F", "G", "H"]           # Bear / Base / Bull on Scenarios tab
SCEN_BASE = "G"
BASE_REV = D.IS['revenue']['FY2025']  # 11,102,600 (blue)

# ------------------------------------------------------------------ COVER
cov = wb.active
cov.title = "Cover"
cov.sheet_view.showGridLines = False
S.set_col_widths(cov, {'A': 3, 'B': 100})
write(cov, 'B2', "GLOBAL INVESTMENT SOCIETY  |  INVESTMENT RESEARCH DIVISION", S.WHITE, bold=True, size=13, fillc=S.DARK)
cov['B2'].alignment = S.Alignment(horizontal='left', vertical='center', indent=1)
cov.row_dimensions[2].height = 26
write(cov, 'B4', "lululemon athletica inc. (NASDAQ: LULU)", S.DARK, bold=True, size=20)
write(cov, 'B5', "Discounted Cash Flow Valuation \u2014 Unlevered Free Cash Flow", S.ACCENT, bold=True, size=13)
write(cov, 'B7', "Recommendation:  LONG / OVERWEIGHT", S.GREEN, bold=True, size=14)
write(cov, 'B9', "FONT / COLOR CONVENTION", S.DARK, bold=True, size=12)
write(cov, 'B10', "Blue font  =  figures reported by the company (click value or source link)", S.BLUE, bold=True, size=11)
write(cov, 'B11', "Black font  =  calculations / formulas", S.BLACK, bold=True, size=11)
write(cov, 'B12', "Red font  =  analyst assumptions — cols B/C/D on every tab: Justification | Source | Ctrl+F", S.RED, bold=True, size=11)
write(cov, 'B14', "TABS", S.DARK, bold=True, size=12)
write(cov, 'B15', "WACC  \u2022  DCF (base case + sensitivity)  \u2022  Scenarios  \u2022  Comps / Football Field", S.BLACK, size=10)
write(cov, 'B17', "SOURCES", S.DARK, bold=True, size=12)
write_link(cov, 'B18', "SEC EDGAR filings, CIK 0001397187 (Form 10-K, FY2025)", D.SOURCES["edgar_xbrl"],
           color=S.BLUE, size=10, hint=D.COVER_HINTS["edgar_xbrl"])
write_link(cov, 'B19', "FY2025 Form 10-K (ended Feb 1, 2026)", D.filing_url("FY2025"),
           color=S.BLUE, size=10, hint=D.COVER_HINTS["filing_fy2025"])
write_link(cov, 'B20', "Market data & Q2 FY2026 results (Sep 3, 2026 earnings release)", D.SOURCES["earnings_sep2026"],
           color=S.BLUE, size=10, hint=D.COVER_HINTS["earnings_sep2026"])
write_link(cov, 'B21', "Current share price (NASDAQ: LULU)", D.SOURCES["nasdaq_quote"],
           color=S.BLUE, size=10, hint=D.COVER_HINTS["nasdaq_quote"])
write(cov, 'B23', "Built from scratch for the GIS IR selection assignment.", S.BLACK, italic=True, size=9)

# ------------------------------------------------------------------ WACC
wacc = wb.create_sheet("WACC")
wacc.sheet_view.showGridLines = False
S.set_col_widths(wacc, {'A': 30, 'B': 28, 'C': 18, 'D': 40, 'E': 12})
write(wacc, 'A1', "WEIGHTED AVERAGE COST OF CAPITAL", S.WHITE, bold=True, size=12, fillc=S.DARK)
for c in ['B', 'C', 'D', 'E']:
    wacc[f'{c}1'].fill = S.fill(S.DARK)
wacc.row_dimensions[1].height = 16
write(wacc, f'{DJ}2', "Justification (~20 words)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(wacc, f'{DS}2', "Source (click)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(wacc, f'{DC}2', "Ctrl+F (prove number)", S.ACCENT, bold=True, size=9, align=S.left_indent)
wacc.freeze_panes = 'E3'
WR = {}
r = [3]
def w_row(key, label, value, color, fmt=PCT, bold=False, top=False, doc_key=None):
    WR[key] = r[0]
    bdr = S.top_border if top else None
    write(wacc, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    if isinstance(value, str):
        write(wacc, f'E{r[0]}', value, S.BLACK, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    else:
        write(wacc, f'E{r[0]}', value, color, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    if doc_key:
        write_assumption_docs(wacc, r[0], DJ, DS, DC, doc_key, D.JUST, D.ASSUMPTION_SRC,
                              hints=D.SOURCE_HINT)
    r[0] += 1

write(wacc, 'A2', "Cost of equity (CAPM)", S.ACCENT, bold=True, size=10)
w_row('rf', "Risk-free rate (10-yr UST)", 0.048, S.RED, doc_key='wacc_rf')
w_row('erp', "Equity risk premium", 0.060, S.RED, doc_key='wacc_erp')
w_row('beta', "Levered beta", 0.95, S.RED, fmt='0.00', doc_key='wacc_beta')
w_row('coe', "Cost of equity = rf + \u03b2 \u00d7 ERP", f"=E{WR['rf']}+E{WR['beta']}*E{WR['erp']}", None, bold=True, top=True)
r[0] += 1
write(wacc, f'A{r[0]}', "Cost of debt", S.ACCENT, bold=True, size=10)
r[0] += 1
w_row('kd', "Pre-tax cost of debt", 0.050, S.RED, doc_key='wacc_kd')
w_row('tax', "Tax rate", 0.300, S.RED, doc_key='wacc_tax')
w_row('kdat', "After-tax cost of debt", f"=E{WR['kd']}*(1-E{WR['tax']})", None, top=True)
r[0] += 1
write(wacc, f'A{r[0]}', "Capital structure (market values)", S.ACCENT, bold=True, size=10)
r[0] += 1
w_row('we', "Equity weight", 1.00, S.RED, doc_key='wacc_we')
w_row('wd', "Debt weight", 0.00, S.RED, doc_key='wacc_wd')
w_row('wacc', "WACC", f"=E{WR['we']}*E{WR['coe']}+E{WR['wd']}*E{WR['kdat']}", None, bold=True, top=True)
wacc[f"E{WR['wacc']}"].font = S.font(color=S.GREEN, bold=True, size=12)
wacc[f"E{WR['wacc']}"].fill = S.fill(S.GREY)

def wref(key):
    return f"WACC!E{WR[key]}"

# ------------------------------------------------------------------ SCENARIOS (built before DCF so base-case drivers can link here)
scn = wb.create_sheet("Scenarios")
scn.sheet_view.showGridLines = False
S.set_col_widths(scn, {'A': 28, 'B': 26, 'C': 16, 'D': 38, 'F': 11, 'G': 11, 'H': 11})
write(scn, 'A1', "SCENARIO ANALYSIS", S.WHITE, bold=True, size=12, fillc=S.DARK)
for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    scn[f'{c}1'].fill = S.fill(S.DARK)
scn.row_dimensions[1].height = 16
write(scn, 'F2', "Bear", S.WHITE, bold=True, size=11, align=S.center, fillc=S.ACCENT)
write(scn, 'G2', "Base", S.WHITE, bold=True, size=11, align=S.center, fillc=S.GREEN)
write(scn, 'H2', "Bull", S.WHITE, bold=True, size=11, align=S.center, fillc=S.DARK)
scn.freeze_panes = 'F3'

SC = {}
rr = [4]
def s_assum(key, label, bear, base, bull, fmt=PCT, doc_key=None, extra_doc_key=None, internal_location=None):
    SC[key] = rr[0]
    write(scn, f'A{rr[0]}', label, S.BLACK, size=10, align=S.left_indent)
    for col, v in zip(SCEN_COLS, [bear, base, bull]):
        write(scn, f'{col}{rr[0]}', v, S.RED, size=10, numfmt=fmt, align=S.right)
    if doc_key:
        write_assumption_docs(scn, rr[0], DJ, DS, DC, doc_key, D.JUST, D.ASSUMPTION_SRC,
                              internal_location=internal_location, hints=D.SOURCE_HINT)
        if extra_doc_key:
            append_assumption_docs(scn, rr[0], DJ, DS, DC, extra_doc_key, D.JUST, D.ASSUMPTION_SRC,
                                   hints=D.SOURCE_HINT, prefix="Also")
        if doc_key in ("sc_capex_pct", "3s_capex_pct") or extra_doc_key == "sc_capex_sales":
            write_ctrl_f(scn, f'{DC}{rr[0]}', D.CAPEX_CTRL_F)
    rr[0] += 1

write(scn, 'A3', "Key assumptions (5-yr forecast)", S.ACCENT, bold=True, size=10)
write(scn, f'{DJ}3', "Justification (~20 words)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(scn, f'{DS}3', "Source (click)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(scn, f'{DC}3', "Ctrl+F (prove number)", S.ACCENT, bold=True, size=9, align=S.left_indent)
s_assum('g1', "FY2026E revenue growth", -0.090, -0.061, -0.040, doc_key='sc_g1')
s_assum('gterm', "FY2027\u2013FY2030E revenue growth (avg)", -0.010, 0.023, 0.060, doc_key='sc_gterm')
s_assum('m1', "Run-rate EBIT margin (clean, ex-refunds)", 0.120, 0.132, 0.145, doc_key='sc_m1')
s_assum('tariff', "FY26 IEEPA tariff refunds ($k, already in guide)",
        D.GUIDANCE['tariff_refund'], D.GUIDANCE['tariff_refund'], D.GUIDANCE['tariff_refund'],
        fmt=NUM, doc_key='sc_tariff')
s_assum('mterm', "Terminal (FY2030E) EBIT margin", 0.120, 0.155, 0.190, doc_key='sc_mterm')
s_assum('wacc', "WACC", 0.115, 0.105, 0.095, doc_key='sc_wacc',
        internal_location=f"'WACC'!E{WR['wacc']}")
s_assum('g', "Terminal growth", 0.015, 0.0225, 0.030, doc_key='sc_g')
s_assum('tax', "Cash tax rate", 0.320, 0.300, 0.280, doc_key='sc_tax')
s_assum('da_pct', "D&A % of revenue", 0.045, 0.045, 0.045, doc_key='sc_da_pct')
s_assum('capex_pct', "Capex % of revenue", 0.070, 0.055, 0.045, doc_key='sc_capex_pct',
        extra_doc_key='sc_capex_sales')
s_assum('ar_pct', "AR % of revenue (NWC)", 0.015, 0.017, 0.020, doc_key='sc_ar')
s_assum('nwc_pct', "Other NWC (ex-AR) % of \u0394revenue", 0.065, 0.058, 0.050, doc_key='sc_nwc_pct')

rr[0] += 1
write(scn, f'A{rr[0]}', "5-year forecast paths (by scenario)", S.ACCENT, bold=True, size=10)
rr[0] += 1

def put(row, lab):
    write(scn, f'A{row}', lab, S.BLACK, size=9, align=S.left_indent)

cur = rr[0]
BASE_AR = D.BS['ar']['FY2025']
rev_rows, mar_rows, ebit_rows, nopat_rows, da_rows, capex_rows = {}, {}, {}, {}, {}, {}
ar_rows, dar_rows, other_nwc_rows, dnwc_rows, fcf_rows = {}, {}, {}, {}, {}
for t in range(1, 6):
    rev_rows[t] = cur
    put(cur, f"  Revenue \u2013 year {t}")
    for col in SCEN_COLS:
        if t == 1:
            f = f"={BASE_REV}*(1+{col}{SC['g1']})"
        else:
            f = f"={col}{rev_rows[t-1]}*(1+{col}{SC['gterm']})"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    mar_rows[t] = cur
    put(cur, f"  Clean EBIT margin \u2013 year {t}")
    for col in SCEN_COLS:
        f = f"={col}{SC['m1']}+({col}{SC['mterm']}-{col}{SC['m1']})*{(t-1)}/4"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=PCT, align=S.right)
    cur += 1
for t in range(1, 6):
    ebit_rows[t] = cur
    put(cur, f"  EBIT \u2013 year {t}")
    for col in SCEN_COLS:
        # FY26 only: add the already-recognized $134.5M refund on top of the clean margin.
        # Later years interpolate 13.2% → 15.5% with no refund leakage.
        extra = f"+{col}{SC['tariff']}" if t == 1 else ""
        f = f"={col}{rev_rows[t]}*{col}{mar_rows[t]}{extra}"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    nopat_rows[t] = cur
    put(cur, f"  NOPAT \u2013 year {t}")
    for col in SCEN_COLS:
        f = f"={col}{ebit_rows[t]}*(1-{col}{SC['tax']})"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    da_rows[t] = cur
    put(cur, f"  D&A \u2013 year {t}")
    for col in SCEN_COLS:
        f = f"={col}{rev_rows[t]}*{col}{SC['da_pct']}"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    capex_rows[t] = cur
    put(cur, f"  Capex \u2013 year {t}")
    for col in SCEN_COLS:
        f = f"={col}{rev_rows[t]}*{col}{SC['capex_pct']}"
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    ar_rows[t] = cur
    put(cur, f"  Accounts receivable \u2013 year {t}")
    for col in SCEN_COLS:
        write(scn, f'{col}{cur}', f"={col}{rev_rows[t]}*{col}{SC['ar_pct']}",
              S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    dar_rows[t] = cur
    put(cur, f"  \u0394AR \u2013 year {t}")
    for col in SCEN_COLS:
        prev_ar = f"{col}{ar_rows[t-1]}" if t > 1 else str(BASE_AR)
        write(scn, f'{col}{cur}', f"={col}{ar_rows[t]}-{prev_ar}",
              S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    other_nwc_rows[t] = cur
    put(cur, f"  Other \u0394NWC (ex-AR) \u2013 year {t}")
    for col in SCEN_COLS:
        prev_rev = f"{col}{rev_rows[t-1]}" if t > 1 else str(BASE_REV)
        write(scn, f'{col}{cur}',
              f"={col}{SC['nwc_pct']}*({col}{rev_rows[t]}-{prev_rev})",
              S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    dnwc_rows[t] = cur
    put(cur, f"  \u0394NWC \u2013 year {t}")
    for col in SCEN_COLS:
        write(scn, f'{col}{cur}', f"={col}{dar_rows[t]}+{col}{other_nwc_rows[t]}",
              S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1
for t in range(1, 6):
    fcf_rows[t] = cur
    put(cur, f"  Unlevered FCF \u2013 year {t}")
    for col in SCEN_COLS:
        f = (f"={col}{nopat_rows[t]}+{col}{da_rows[t]}"
             f"-{col}{capex_rows[t]}-{col}{dnwc_rows[t]}")
        write(scn, f'{col}{cur}', f, S.BLACK, size=9, numfmt=NUM, align=S.right)
    cur += 1

cur += 1
write(scn, f'A{cur}', "Enterprise value (DCF)", S.DARK, bold=True, size=10, align=S.left_indent)
ev_row = cur
for col in SCEN_COLS:
    fcf_cells = ",".join(f"{col}{fcf_rows[t]}" for t in range(1, 6))
    lastf = f"{col}{fcf_rows[5]}"
    f = (f"=NPV({col}{SC['wacc']},{fcf_cells})"
         f"+({lastf}*(1+{col}{SC['g']})/({col}{SC['wacc']}-{col}{SC['g']}))/(1+{col}{SC['wacc']})^5")
    write(scn, f'{col}{cur}', f, S.BLACK, bold=True, size=10, numfmt=NUM, align=S.right, bdr=S.top_border)
cur += 1
write(scn, f'A{cur}', "Implied share price", S.DARK, bold=True, size=11, align=S.left_indent)
pt_row = cur
for col in SCEN_COLS:
    f = f"=({col}{ev_row}+{D.MKT['cash']}-{D.MKT['debt']})/{D.MKT['shares_out']}"
    color = S.GREEN if col == SCEN_BASE else S.BLACK
    write(scn, f'{col}{cur}', f, color, bold=True, size=12, numfmt=MONEY, align=S.right, bdr=S.top_double, fillc=S.GREY)
cur += 1
write(scn, f'A{cur}', "Upside / (downside) vs current", S.DARK, bold=True, size=10, align=S.left_indent)
for col in SCEN_COLS:
    write(scn, f'{col}{cur}', f"={col}{pt_row}/{D.MKT['price']}-1", S.BLACK, bold=True, size=10, numfmt=PCT, align=S.right)
cur += 2
write(scn, f'A{cur}', "Current price $%.2f; cash $%s k; net debt $0 (net-cash balance sheet)." % (D.MKT['price'], f"{D.MKT['cash']:,}"),
      S.BLACK, italic=True, size=8, align=S.left_indent)

def bref(key):
    return f"Scenarios!${SCEN_BASE}${SC[key]}"

YEAR_MAP = {"F": 1, "G": 2, "H": 3, "I": 4, "J": 5}

# ------------------------------------------------------------------ DCF (base — linked to Scenarios → Base column G)
dcf = wb.create_sheet("DCF")
dcf.sheet_view.showGridLines = False
S.set_col_widths(dcf, {'A': 30, 'B': 26, 'C': 16, 'D': 36, 'E': 12, 'F': 11, 'G': 11, 'H': 11, 'I': 11, 'J': 11, 'K': 8, 'L': 16, 'M': 36})
write(dcf, 'A1', "DISCOUNTED CASH FLOW \u2014 BASE CASE  (US$ thousands)", S.WHITE, bold=True, size=12, fillc=S.DARK)
for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    dcf[f'{c}1'].fill = S.fill(S.DARK)
write(dcf, f'{FY25_COL}2', "FY2025A", S.BLUE, bold=True, size=10, align=S.center)
write_link(dcf, f'{FY25_COL}3', "10-K", D.filing_url("FY2025"), color=S.BLUE, size=8, align=S.center, italic=True)
write_ctrl_f(dcf, f'{DS}3', D.COVER_HINTS["filing_fy2025"])
for y in FY:
    write(dcf, f'{FCOL[y]}2', y, S.WHITE, bold=True, size=10, align=S.center, fillc=S.ACCENT)
write(dcf, 'A3', "Forecast drivers linked to Scenarios tab \u2192 Base case (column G)", S.GREY, italic=True, size=9, align=S.left_indent)
write(dcf, f'{CHK_COL}2', "\u0394 vs Scenarios", S.ACCENT, bold=True, size=8, align=S.center)
write(dcf, f'{DJ}2', "Justification (~20 words)", S.ACCENT, bold=True, size=8, align=S.left_indent)
write(dcf, f'{DS}2', "Source (click)", S.ACCENT, bold=True, size=8, align=S.left_indent)
write(dcf, f'{DC}2', "Ctrl+F (prove number)", S.ACCENT, bold=True, size=8, align=S.left_indent)
write(dcf, 'L2', "Alt. source", S.ACCENT, bold=True, size=8, align=S.left_indent)
write(dcf, 'M2', "Alt. Ctrl+F", S.ACCENT, bold=True, size=8, align=S.left_indent)
write(dcf, f'{CHK_COL}3', "(should be 0)", S.GREY, italic=True, size=7, align=S.center)
dcf.freeze_panes = f'{FY25_COL}3'
dcf.row_dimensions[1].height = 16

DR = {}
r = [5]

def _sc_status(row_num, sc_rows, fmt=NUM):
    tol = "0.0001" if fmt == PCT else "0.5"
    parts = [f"ABS({c}{row_num}-Scenarios!${SCEN_BASE}${sc_rows[YEAR_MAP[c]]})" for c in FCOLS]
    return f'=IF(MAX({",".join(parts)})<{tol},"OK","CHECK")'

def d_row(key, label, cval, proj_fn, color_c=S.BLUE, color_p=S.BLACK, fmt=NUM, bold=False,
          top=False, dbl=False, red=False, sc_rows=None, sc_sign=1, justify_key="", extra_doc_key=""):
    DR[key] = r[0]
    row_num = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(dcf, f'A{row_num}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    if cval is not None:
        write_reported(dcf, f'{FY25_COL}{row_num}', cval, D.filing_url("FY2025"), bold=bold, size=10, numfmt=fmt, bdr=bdr)
    for y in FY:
        c = FCOL[y]
        col = S.RED if red else color_p
        write(dcf, f'{c}{row_num}', proj_fn(c), col, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    if sc_rows:
        if sc_sign == -1:
            tol = "0.5"
            parts = [f"ABS({c}{row_num}+Scenarios!${SCEN_BASE}${sc_rows[YEAR_MAP[c]]})" for c in FCOLS]
            status = f'=IF(MAX({",".join(parts)})<{tol},"OK","CHECK")'
        else:
            status = _sc_status(row_num, sc_rows, fmt)
        write(dcf, f'{CHK_COL}{row_num}', status, S.BLACK, bold=bold, size=8, align=S.center)
    if justify_key:
        write_assumption_docs(dcf, row_num, DJ, DS, DC, justify_key, D.JUST, D.ASSUMPTION_SRC,
                              hints=D.SOURCE_HINT)
        if extra_doc_key:
            append_assumption_docs(dcf, row_num, DJ, DS, DC, extra_doc_key, D.JUST, D.ASSUMPTION_SRC,
                                   hints=D.SOURCE_HINT, prefix="Also")
        if justify_key in ("sc_capex_pct", "3s_capex_pct") or extra_doc_key == "sc_capex_sales":
            write_ctrl_f(dcf, f'{DC}{row_num}', D.CAPEX_CTRL_F)
    elif extra_doc_key:
        src = D.ASSUMPTION_SRC.get(extra_doc_key)
        if src and src[1]:
            write_link(dcf, f'L{row_num}', src[0], src[1], color=S.BLUE, size=8, italic=True)
            write_ctrl_f(dcf, f'M{row_num}', D.SOURCE_HINT.get(extra_doc_key))
    r[0] += 1

PREVF = {"F": FY25_COL, "G": "F", "H": "G", "I": "H", "J": "I"}

# Forecast rows — each line links to Scenarios base (column G)
d_row('rev', "Net revenue", BASE_REV,
      lambda c: f"=Scenarios!{SCEN_BASE}{rev_rows[YEAR_MAP[c]]}", sc_rows=rev_rows)
d_row('growth', "Revenue growth %", None,
      lambda c: f"={c}{DR['rev']}/{PREVF[c]}{DR['rev']}-1", fmt=PCT,
      justify_key="sc_g1", extra_doc_key="sc_gterm")
d_row('margin', "Clean / run-rate EBIT margin %", None,
      lambda c: f"=Scenarios!{SCEN_BASE}{mar_rows[YEAR_MAP[c]]}", fmt=PCT, sc_rows=mar_rows,
      justify_key="sc_m1", extra_doc_key="sc_mterm")
d_row('tariff', "Plus: FY26 IEEPA tariff refunds (one-time)", 0,
      lambda c: f"={bref('tariff')}" if YEAR_MAP[c] == 1 else 0,
      fmt=NUM, red=True, justify_key="sc_tariff")
d_row('ebit', "EBIT (incl. FY26 refund)", D.IS['operating_income']['FY2025'],
      lambda c: f"=Scenarios!{SCEN_BASE}{ebit_rows[YEAR_MAP[c]]}", color_c=S.BLUE, bold=True, top=True,
      sc_rows=ebit_rows)
d_row('rep_margin', "Reported EBIT margin % (incl. FY26 refund)",
      D.IS['operating_income']['FY2025'] / BASE_REV,
      lambda c: f"={c}{DR['ebit']}/{c}{DR['rev']}", fmt=PCT, color_c=S.BLUE)
d_row('taxes', "Less: cash taxes on EBIT", None,
      lambda c: f"=-Scenarios!{SCEN_BASE}{ebit_rows[YEAR_MAP[c]]}*{bref('tax')}")
d_row('nopat', "NOPAT", None,
      lambda c: f"=Scenarios!{SCEN_BASE}{nopat_rows[YEAR_MAP[c]]}", bold=True, top=True, sc_rows=nopat_rows)
d_row('da_pct', "  D&A % of revenue", None, lambda c: f"={bref('da_pct')}", fmt=PCT, red=True,
      justify_key="sc_da_pct")
d_row('da', "Plus: depreciation & amortization", None,
      lambda c: f"=Scenarios!{SCEN_BASE}{da_rows[YEAR_MAP[c]]}", sc_rows=da_rows)
d_row('capex_pct', "  Capex % of revenue", None, lambda c: f"={bref('capex_pct')}", fmt=PCT, red=True,
      justify_key="sc_capex_pct", extra_doc_key="sc_capex_sales")
d_row('capex', "Less: capital expenditures", None,
      lambda c: f"=-Scenarios!{SCEN_BASE}{capex_rows[YEAR_MAP[c]]}", sc_rows=capex_rows, sc_sign=-1)
d_row('ar_pct', "  AR % of revenue", None, lambda c: f"={bref('ar_pct')}", fmt=PCT, red=True,
      justify_key="sc_ar")
d_row('ar', "Accounts receivable, net", BASE_AR,
      lambda c: f"=Scenarios!{SCEN_BASE}{ar_rows[YEAR_MAP[c]]}", color_c=S.BLUE, sc_rows=ar_rows,
      justify_key="sc_ar")
d_row('dar', "  \u0394AR (increase)/decrease \u2014 part of NWC", None,
      lambda c: f"=-Scenarios!{SCEN_BASE}{dar_rows[YEAR_MAP[c]]}", sc_rows=dar_rows, sc_sign=-1)
d_row('nwc_pct', "  Other NWC (ex-AR) % of \u0394revenue", None, lambda c: f"={bref('nwc_pct')}", fmt=PCT, red=True,
      justify_key="sc_nwc_pct")
d_row('other_nwc', "  Other \u0394NWC (ex-AR) (increase)/decrease", None,
      lambda c: f"=-Scenarios!{SCEN_BASE}{other_nwc_rows[YEAR_MAP[c]]}",
      sc_rows=other_nwc_rows, sc_sign=-1)
d_row('dnwc', "Total (increase)/decrease in net working capital", None,
      lambda c: f"=-Scenarios!{SCEN_BASE}{dnwc_rows[YEAR_MAP[c]]}", sc_rows=dnwc_rows, sc_sign=-1)
write(dcf, f'A{r[0]}',
      "  memo: \u0394NWC = \u0394AR + other \u0394NWC. AR is 1.7% of sales (own line); other is 5.8% of \u0394sales.",
      S.BLACK, italic=True, size=8, align=S.left_indent)
r[0] += 1
d_row('ufcf', "Unlevered free cash flow", None,
      lambda c: f"=Scenarios!{SCEN_BASE}{fcf_rows[YEAR_MAP[c]]}", bold=True, top=True, dbl=True, sc_rows=fcf_rows)
d_row('period', "Discount period (years)", None, lambda c: {"F": 1, "G": 2, "H": 3, "I": 4, "J": 5}[c], fmt='0')
d_row('df', "Discount factor @ WACC", None, lambda c: f"=1/(1+{bref('wacc')})^{c}{DR['period']}", fmt='0.000')
d_row('pv', "PV of unlevered FCF", None, lambda c: f"={c}{DR['ufcf']}*{c}{DR['df']}", bold=True, top=True)

r[0] += 1
write(dcf, f'A{r[0]}', f"Scenarios linkage summary (column {CHK_COL} should all read OK)", S.ACCENT, bold=True, size=9, align=S.left_indent)
chk_start, chk_end = DR['rev'], DR['ufcf']
write(dcf, f'{CHK_COL}{r[0]}',
      f'=IF(COUNTIF({CHK_COL}{chk_start}:{CHK_COL}{chk_end},"CHECK")=0,"ALL OK","REVIEW")',
      S.GREEN, bold=True, size=9, align=S.center)
r[0] += 1
# Valuation block
VR = {}
def v_row(key, label, formula, fmt=NUM, color=S.BLACK, bold=False, top=False, dbl=False, red=False,
          source_url=None, source_label="", source_hint=None, doc_key=None, internal_location=None):
    VR[key] = r[0]
    bdr = S.top_double if dbl else (S.top_border if top else None)
    write(dcf, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    col = S.RED if red else color
    if isinstance(formula, (int, float)):
        if source_url and color == S.BLUE:
            write_reported(dcf, f'{FY25_COL}{r[0]}', formula, source_url, bold=bold, size=10, numfmt=fmt, bdr=bdr)
        else:
            write(dcf, f'{FY25_COL}{r[0]}', formula, col, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    else:
        write(dcf, f'{FY25_COL}{r[0]}', formula, col, bold=bold, size=10, numfmt=fmt, align=S.right, bdr=bdr)
    if source_url and source_label:
        write_link(dcf, f'{DS}{r[0]}', source_label, source_url, color=S.BLUE, size=8, italic=True)
        write_ctrl_f(dcf, f'{DC}{r[0]}', source_hint)
    elif source_label:
        write(dcf, f'{DS}{r[0]}', source_label, S.BLACK, italic=True, size=8, align=S.left_indent)
    if doc_key:
        write_assumption_docs(dcf, r[0], DJ, DS, DC, doc_key, D.JUST, D.ASSUMPTION_SRC,
                              internal_location=internal_location, hints=D.SOURCE_HINT)
    r[0] += 1

write(dcf, f'A{r[0]}', "VALUATION \u2014 GORDON GROWTH (PERPETUITY) METHOD", S.WHITE, bold=True, size=10, fillc=S.DARK)
for c in ['B', 'C']:
    dcf[f'{c}{r[0]}'].fill = S.fill(S.DARK)
r[0] += 1
v_row('sumpv', "Sum of PV of explicit FCF (FY26\u2013FY30)", f"=SUM(F{DR['pv']}:J{DR['pv']})", bold=True)
v_row('g', "Terminal growth rate (g)", f"={bref('g')}", fmt=PCT, doc_key='sc_g')
v_row('ebitda', "Terminal EBITDA (FY2030E EBIT + D&A)", f"=J{DR['ebit']}+J{DR['da']}", bold=True, top=True)
v_row('tv', "Terminal value = FCF\u2085\u00d7(1+g)/(WACC\u2212g)",
      f"=J{DR['ufcf']}*(1+{bref('g')})/({bref('wacc')}-{bref('g')})")
v_row('implied_exit', "  Implied exit EV/EBITDA (Gordon Growth)",
      f"=E{VR['tv']}/E{VR['ebitda']}", fmt=MULT)
v_row('pvtv', "PV of terminal value", f"=E{VR['tv']}/(1+{bref('wacc')})^J{DR['period']}", bold=True)
v_row('ev', "Enterprise value", f"=E{VR['sumpv']}+E{VR['pvtv']}", bold=True, top=True)
v_row('cash', "Plus: cash & equivalents (FY2025)", D.MKT['cash'], color=S.BLUE,
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_bs"])
v_row('debt', "Less: total debt", -D.MKT['debt'], color=S.BLUE,
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_debt"])
v_row('eqv', "Equity value", f"=E{VR['ev']}+E{VR['cash']}+E{VR['debt']}", bold=True, top=True)
v_row('sh', "Diluted shares outstanding (000)", D.MKT['shares_out'], color=S.BLUE,
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_shares"])
v_row('pt', "Implied value per share", f"=E{VR['eqv']}/E{VR['sh']}", fmt=MONEY, bold=True, top=True, dbl=True)
dcf[f"E{VR['pt']}"].font = S.font(color=S.GREEN, bold=True, size=13)
dcf[f"E{VR['pt']}"].fill = S.fill(S.GREY)
write(dcf, f'{CHK_COL}{VR["pt"]}', f'=IF(ABS(E{VR["pt"]}-Scenarios!${SCEN_BASE}${pt_row})<0.05,"OK","CHECK")',
      S.BLACK, bold=True, size=8, align=S.center)
v_row('px', "Current share price", D.MKT['price'], fmt=MONEY, color=S.BLUE,
      source_url=D.SOURCES["nasdaq_quote"], source_label="NASDAQ", source_hint=D.REPORTED_HINTS["nasdaq"])
v_row('upside', "Implied upside / (downside)", f"=E{VR['pt']}/E{VR['px']}-1", fmt=PCT, bold=True)
dcf[f"E{VR['upside']}"].font = S.font(color=S.GREEN, bold=True, size=11)
v_row('tvpct', "  memo: % of EV from terminal value", f"=E{VR['pvtv']}/E{VR['ev']}", fmt=PCT)

r[0] += 1
write(dcf, f'A{r[0]}', "CROSS-CHECK \u2014 EXIT MULTIPLE METHOD", S.WHITE, bold=True, size=10, fillc=S.DARK)
for c in ['B', 'C']:
    dcf[f'{c}{r[0]}'].fill = S.fill(S.DARK)
r[0] += 1
write(dcf, f'A{r[0]}', "Exit multiple selection (see Comps tab for peer build)", S.ACCENT, bold=True, size=9, align=S.left_indent)
r[0] += 1
v_row('exitm', "Selected exit EV/EBITDA multiple (FY2030E)", 8.0, fmt=MULT, red=True, doc_key='dcf_exitm')
v_row('tv2', "Terminal value = Terminal EBITDA \u00d7 exit multiple", f"=E{VR['ebitda']}*E{VR['exitm']}")
v_row('pvtv2', "PV of terminal value", f"=E{VR['tv2']}/(1+{bref('wacc')})^J{DR['period']}")
v_row('ev2', "Enterprise value (exit method)", f"=E{VR['sumpv']}+E{VR['pvtv2']}", bold=True, top=True)
v_row('pt2', "Implied value per share (exit method)",
      f"=(E{VR['ev2']}+E{VR['cash']}+E{VR['debt']})/E{VR['sh']}", fmt=MONEY, bold=True)
dcf[f"E{VR['pt2']}"].font = S.font(color=S.GREEN, bold=True, size=11)

r[0] += 1
write(dcf, f'A{r[0]}', "TERMINAL VALUE RECONCILIATION (GORDON vs EXIT MULTIPLE)", S.WHITE, bold=True, size=10, fillc=S.DARK)
for c in ['B', 'C', 'D']:
    dcf[f'{c}{r[0]}'].fill = S.fill(S.DARK)
r[0] += 1
write(dcf, f'A{r[0]}', "", S.BLACK, size=9)
write(dcf, f'B{r[0]}', "Gordon Growth", S.WHITE, bold=True, size=9, align=S.center, fillc=S.NAVY)
write(dcf, f'C{r[0]}', "Exit Multiple", S.WHITE, bold=True, size=9, align=S.center, fillc=S.NAVY)
write(dcf, f'D{r[0]}', "Variance", S.WHITE, bold=True, size=9, align=S.center, fillc=S.NAVY)
r[0] += 1

def recon_row(label, gordon, exit_, variance, fmt=NUM, bold=False):
    write(dcf, f'A{r[0]}', label, S.DARK if bold else S.BLACK, bold=bold, size=10, align=S.left_indent)
    write(dcf, f'B{r[0]}', gordon, S.BLACK, bold=bold, size=10, numfmt=fmt, align=S.right)
    write(dcf, f'C{r[0]}', exit_, S.BLACK, bold=bold, size=10, numfmt=fmt, align=S.right)
    write(dcf, f'D{r[0]}', variance, S.BLACK, bold=bold, size=10, numfmt=fmt, align=S.right)
    r[0] += 1

recon_row("Terminal value ($)",
          f"=E{VR['tv']}", f"=E{VR['tv2']}", f"=B{r[0]}-C{r[0]}", bold=True)
tv_var_row = r[0] - 1
recon_row("Exit EV/EBITDA (implied vs selected)",
          f"=E{VR['implied_exit']}", f"=E{VR['exitm']}", f"=B{r[0]}-C{r[0]}", fmt=MULT, bold=True)
mult_var_row = r[0] - 1
recon_row("Terminal value variance (%)",
          "", "", f"=(B{tv_var_row}-C{tv_var_row})/B{tv_var_row}", fmt=PCT)
recon_row("Implied share price",
          f"=E{VR['pt']}", f"=E{VR['pt2']}", f"=B{r[0]}-C{r[0]}", fmt=MONEY, bold=True)
# sanity check flag: multiples within 1.5 turns
write(dcf, f'A{r[0]}', "Sanity check: multiples within \u00b11.5 turns?", S.BLACK, bold=True, size=10, align=S.left_indent)
write(dcf, f'B{r[0]}', f'=IF(ABS(D{mult_var_row})<=1.5,"PASS","REVIEW")', S.GREEN, bold=True, size=11, align=S.center)
write(dcf, f'C{r[0]}', f'=TEXT(D{mult_var_row},"0.0")&"x spread vs 8.0x exit multiple"', S.BLACK, italic=True, size=9, align=S.left_indent)
write(dcf, f'D{r[0]}', "Gordon implied should bracket exit assumption", S.BLACK, italic=True, size=8, align=S.left_indent)
r[0] += 1

# ------------------------------------------------------------------ SENSITIVITY (WACC x g)
r[0] += 2
write(dcf, f'A{r[0]}', "SENSITIVITY \u2014 IMPLIED SHARE PRICE (WACC vs terminal growth)", S.WHITE, bold=True, size=10, fillc=S.DARK)
for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    dcf[f'{c}{r[0]}'].fill = S.fill(S.DARK)
r[0] += 1
sens_top = r[0]
waccs = [0.095, 0.100, 0.105, 0.110, 0.115]
gs = [0.015, 0.020, 0.0225, 0.025, 0.030]
write(dcf, f'A{sens_top}', "WACC \\ g", S.DARK, bold=True, size=9, align=S.center, fillc=S.LIGHT)
gcols = FCOLS
for j, g in enumerate(gs):
    write(dcf, f'{gcols[j]}{sens_top}', g, S.RED, bold=True, size=9, numfmt=PCT, align=S.center, fillc=S.LIGHT)
write_assumption_docs(dcf, sens_top, DJ, DS, DC, 'sens_g', D.JUST, D.ASSUMPTION_SRC,
                      hints=D.SOURCE_HINT)
write_internal_link(dcf, f'L{sens_top}', 'Scenarios: terminal g', f"'Scenarios'!{SCEN_BASE}{SC['g']}")
fcf_rng = f"F{DR['ufcf']}:J{DR['ufcf']}"
lastfcf = f"J{DR['ufcf']}"
for i, wv in enumerate(waccs):
    rr = sens_top + 1 + i
    write(dcf, f'A{rr}', wv, S.RED, bold=True, size=9, numfmt=PCT, align=S.center, fillc=S.LIGHT)
    write_internal_link(dcf, f'L{rr}', 'WACC tab', f"'WACC'!E{WR['wacc']}")
    if i == 0:
        src = D.ASSUMPTION_SRC['sens_wacc']
        write_link(dcf, f'{DS}{rr}', src[0], src[1], color=S.BLUE, size=8, italic=True)
    write_ctrl_f(dcf, f'{DC}{rr}', D.SOURCE_HINT['sens_wacc'])
    for j, g in enumerate(gs):
        f = (f"=(NPV({wv},{fcf_rng})"
             f"+({lastfcf}*(1+{g})/({wv}-{g}))/(1+{wv})^5"
             f"+E{VR['cash']}+E{VR['debt']})/E{VR['sh']}")
        col = S.GREEN if (abs(wv-0.105) < 1e-9 and abs(g-0.0225) < 1e-9) else S.BLACK
        write(dcf, f'{gcols[j]}{rr}', f, col, size=9, numfmt=MONEY, align=S.center)
r[0] = sens_top + 1 + len(waccs)
write_assumption_docs(dcf, r[0], DJ, DS, DC, 'sens_axes', D.JUST, D.ASSUMPTION_SRC,
                      internal_location=f"'Scenarios'!{SCEN_BASE}{SC['wacc']}",
                      extra_source_col='L', extra_ctrl_f_col='M',
                      extra_label='FRED: Real GDP (GDPC1)',
                      extra_url=D.SOURCES['fred_gdpc1'], extra_hint=D.SOURCE_HINT['sens_g'],
                      hints=D.SOURCE_HINT)
r[0] += 1

# ------------------------------------------------------------------ COMPS / FOOTBALL FIELD
comps = wb.create_sheet("Comps")
comps.sheet_view.showGridLines = False
S.set_col_widths(comps, {'A': 34, 'B': 26, 'C': 16, 'D': 40, 'E': 14, 'F': 14, 'G': 12, 'H': 12})
write(comps, 'A1', "RELATIVE VALUATION \u2014 IMPLIED PRICE RANGES (FOOTBALL FIELD)", S.WHITE, bold=True, size=12, fillc=S.DARK)
for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    comps[f'{c}1'].fill = S.fill(S.DARK)
comps.row_dimensions[1].height = 16
write(comps, f'{DJ}2', "Justification (~20 words)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(comps, f'{DS}2', "Source (click)", S.ACCENT, bold=True, size=9, align=S.left_indent)
write(comps, f'{DC}2', "Ctrl+F (prove number)", S.ACCENT, bold=True, size=9, align=S.left_indent)
comps.freeze_panes = 'E3'
write(comps, 'A3', "LULU operating metrics (FY2025A / FY2026E)", S.ACCENT, bold=True, size=10)
CM = {}
rr = [4]
def c_row(key, label, val, color=S.BLUE, fmt=NUM, source_url=None, source_label="", source_hint=None):
    CM[key] = rr[0]
    write(comps, f'A{rr[0]}', label, S.BLACK, size=10, align=S.left_indent)
    if source_url and color == S.BLUE and isinstance(val, (int, float)):
        write_reported(comps, f'E{rr[0]}', val, source_url, size=10, numfmt=fmt)
    else:
        write(comps, f'E{rr[0]}', val, color, size=10, numfmt=fmt, align=S.right)
    if source_url and source_label:
        write_source_with_ctrl_f(comps, f'{DS}{rr[0]}', f'{DC}{rr[0]}', source_label, source_url, source_hint)
    rr[0] += 1

c_row('rev', "FY2025A revenue", D.IS['revenue']['FY2025'],
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_is"])
c_row('ebitda', "FY2025A EBITDA (EBIT + D&A)",
      D.IS['operating_income']['FY2025'] + D.CF['d_and_a']['FY2025'])
comps[f"E{CM['ebitda']}"].value = f"={D.IS['operating_income']['FY2025']}+{D.CF['d_and_a']['FY2025']}"
comps[f"E{CM['ebitda']}"].font = S.font(color=S.BLACK)
CM['ebitda30'] = rr[0]
write(comps, f'A{rr[0]}', "FY2030E terminal EBITDA (DCF EBIT + D&A)", S.BLACK, size=10, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=DCF!E{VR['ebitda']}", S.BLACK, size=10, numfmt=NUM, align=S.right)
write_internal_link(comps, f'{DS}{rr[0]}', "↳ DCF base case", f"'DCF'!E{VR['ebitda']}")
rr[0] += 1
c_row('eps26', "FY2026E diluted EPS (guidance midpoint)", 9.61, color=S.BLUE, fmt=EPSFMT,
      source_url=D.SOURCES["earnings_sep2026"], source_label="Release",
      source_hint=D.REPORTED_HINTS["earnings_eps"])
c_row('cash', "Cash & equivalents", D.MKT['cash'],
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_bs"])
c_row('sh', "Diluted shares (000)", D.MKT['shares_out'],
      source_url=D.filing_url("FY2025"), source_label="10-K", source_hint=D.REPORTED_HINTS["10k_shares"])
c_row('px', "Current share price", D.MKT['price'], fmt=MONEY,
      source_url=D.SOURCES["nasdaq_quote"], source_label="NASDAQ",
      source_hint=D.REPORTED_HINTS["nasdaq"])
# current multiples (black formulas)
CM['ceved'] = rr[0]
write(comps, f'A{rr[0]}', "  memo: current EV / EBITDA", S.BLACK, italic=True, size=9, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=(E{CM['px']}*E{CM['sh']}-E{CM['cash']})/E{CM['ebitda']}", S.BLACK, italic=True, size=9, numfmt=MULT, align=S.right)
rr[0] += 1
write(comps, f'A{rr[0]}', "  memo: current P / E (FY2026E)", S.BLACK, italic=True, size=9, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=E{CM['px']}/E{CM['eps26']}", S.BLACK, italic=True, size=9, numfmt=MULT, align=S.right)
rr[0] += 2

# ---- Exit multiple build (supports DCF 8.0x terminal assumption) ----
write(comps, f'A{rr[0]}', "EXIT MULTIPLE BUILD \u2014 FY2030E TERMINAL YEAR", S.ACCENT, bold=True, size=10)
rr[0] += 1
write(comps, f'A{rr[0]}', "Peer / reference EV/EBITDA (forward / illustrative)", S.BLACK, italic=True, size=9, align=S.left_indent)
rr[0] += 1
write(comps, f'A{rr[0]}', "Company", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DJ}{rr[0]}', "Justification", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DS}{rr[0]}', "Source (click)", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DC}{rr[0]}', "Ctrl+F (prove number)", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'E{rr[0]}', "EV/EBITDA", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.center)
comps.freeze_panes = 'E4'
rr[0] += 1
peer_rows = [
    ("lululemon (LULU) \u2014 current", f"=(E{CM['px']}*E{CM['sh']}-E{CM['cash']})/E{CM['ebitda']}", True,
     "Distressed trough multiple on FY2025A EBITDA", D.filing_url("FY2025"), "LULU FY2025 10-K"),
    ("Nike (NKE)", 12.0, False, "comps_nke"),
    ("Deckers (DECK)", 8.0, False, "comps_deck"),
    ("On Holding (ONON)", 14.0, False, "comps_onon"),
    ("adidas (ADS)", 9.3, False, "comps_ads"),
    ("V.F. Corp (VFC)", 10.7, False, "comps_vfc"),
]
EM = {}
for row in peer_rows:
    if len(row) == 6:
        name, mult, is_formula, just_text, src_url, src_label = row
        src_key = None
    else:
        name, mult, is_formula, src_key = row
        just_text = src_url = src_label = None
    EM[name] = rr[0]
    write(comps, f'A{rr[0]}', name, S.BLACK, size=10, align=S.left_indent)
    if is_formula:
        write(comps, f'E{rr[0]}', mult, S.BLACK, size=10, numfmt=MULT, align=S.center)
        if src_url:
            write_link(comps, f'{DS}{rr[0]}', src_label, src_url, color=S.BLUE, size=8, italic=True)
            write_ctrl_f(comps, f'{DC}{rr[0]}', D.REPORTED_HINTS["10k_ebitda"])
        write(comps, f'{DJ}{rr[0]}', just_text, S.BLACK, italic=True, size=8, align=S.left_indent)
    else:
        write(comps, f'E{rr[0]}', mult, S.RED, size=10, numfmt=MULT, align=S.center)
        write_assumption_docs(comps, rr[0], DJ, DS, DC, src_key, D.JUST, D.ASSUMPTION_SRC,
                              hints=D.SOURCE_HINT)
    rr[0] += 1
rr[0] += 1
write(comps, f'A{rr[0]}', "Terminal multiple selection (DCF exit method)", S.ACCENT, bold=True, size=10)
rr[0] += 1
EM['gordon'] = rr[0]
write(comps, f'A{rr[0]}', "Gordon Growth implied exit EV/EBITDA", S.BLACK, size=10, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=DCF!E{VR['implied_exit']}", S.BLACK, size=10, numfmt=MULT, align=S.right)
rr[0] += 1
EM['selected'] = rr[0]
write(comps, f'A{rr[0]}', "Selected exit multiple (base case)", S.BLACK, bold=True, size=10, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=DCF!E{VR['exitm']}", S.BLACK, bold=True, size=10, numfmt=MULT, align=S.right)
rr[0] += 1
write(comps, f'A{rr[0]}', "Spread (selected \u2212 Gordon implied)", S.BLACK, size=10, align=S.left_indent)
write(comps, f'E{rr[0]}', f"=E{EM['selected']}-E{EM['gordon']}", S.BLACK, size=10, numfmt=MULT, align=S.right)
rr[0] += 1
write(comps, f'A{rr[0]}', "Rationale for 8.0x", S.BLACK, bold=True, size=10, align=S.left_indent)
rr[0] += 1
for bullet in [
    "\u2022  Mid-point of terminal football field (6.5\u20139.5x on FY2030E EBITDA) \u2014 8.0x sits between bear and bull exit",
    "\u2022  Above Gordon-implied exit (~7x) \u2014 ~1 turn buffer vs perpetuity math on terminal FCF",
    "\u2022  Below premium-growth peers (ONON ~25x, DECK ~15x) \u2014 reflects Americas maturity at terminal",
    "\u2022  Above current LULU (~3.5x on FY2025A) \u2014 assumes partial recovery, not full re-rating to historical peaks",
    "\u2022  Peer simple avg ~16x not used \u2014 terminal multiple discounted for lower terminal growth vs ONON/NKE",
]:
    write(comps, f'A{rr[0]}', bullet, S.BLACK, size=9, align=S.left_indent)
    rr[0] += 1
rr[0] += 1

write(comps, f'A{rr[0]}', "Methodology / multiple ranges (FY2030E terminal EBITDA \u2014 football field)", S.ACCENT, bold=True, size=10)
rr[0] += 1
write(comps, f'A{rr[0]}', "Method", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DJ}{rr[0]}', "Justification", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DS}{rr[0]}', "Source", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'{DC}{rr[0]}', "Ctrl+F", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.left_indent)
write(comps, f'E{rr[0]}', "Low mult.", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.center)
write(comps, f'F{rr[0]}', "High mult.", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.center)
write(comps, f'G{rr[0]}', "Implied px (low)", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.center)
write(comps, f'H{rr[0]}', "Implied px (high)", S.WHITE, bold=True, size=10, fillc=S.DARK, align=S.center)
rr[0] += 1

def ff_ev_ebitda(label, lo, hi, ebitda_key='ebitda30'):
    write(comps, f'A{rr[0]}', label, S.BLACK, size=10, align=S.left_indent)
    write(comps, f'E{rr[0]}', lo, S.RED, size=10, numfmt=MULT, align=S.center)
    write(comps, f'F{rr[0]}', hi, S.RED, size=10, numfmt=MULT, align=S.center)
    for outcol, mcol in [('G', 'E'), ('H', 'F')]:
        f = f"=(E{CM[ebitda_key]}*{mcol}{rr[0]}+E{CM['cash']})/E{CM['sh']}"
        write(comps, f'{outcol}{rr[0]}', f, S.BLACK, size=10, numfmt=MONEY, align=S.center)
    write_ff_dual_docs(comps, rr[0], 'comps_ff_ev_lo', 'comps_ff_ev_hi', DJ, DS, DC,
                       D.JUST, D.ASSUMPTION_SRC, hints=D.SOURCE_HINT)
    rr[0] += 1

def ff_pe(label, lo, hi):
    write(comps, f'A{rr[0]}', label, S.BLACK, size=10, align=S.left_indent)
    write(comps, f'E{rr[0]}', lo, S.RED, size=10, numfmt=MULT, align=S.center)
    write(comps, f'F{rr[0]}', hi, S.RED, size=10, numfmt=MULT, align=S.center)
    for outcol, mcol in [('G', 'E'), ('H', 'F')]:
        f = f"=E{CM['eps26']}*{mcol}{rr[0]}"
        write(comps, f'{outcol}{rr[0]}', f, S.BLACK, size=10, numfmt=MONEY, align=S.center)
    write_ff_dual_docs(comps, rr[0], 'comps_ff_pe_lo', 'comps_ff_pe_hi', DJ, DS, DC,
                       D.JUST, D.ASSUMPTION_SRC, hints=D.SOURCE_HINT)
    rr[0] += 1

ff_ev_ebitda("EV / EBITDA (FY2030E terminal)", 6.5, 9.5)
# DCF exit-method row (8.0x on same terminal EBITDA base)
exit_ff_row = rr[0]
write(comps, f'A{exit_ff_row}', "DCF exit method (8.0x on FY2030E EBITDA)", S.BLACK, bold=True, size=10, align=S.left_indent)
write(comps, f'E{exit_ff_row}', f"=DCF!E{VR['exitm']}", S.BLACK, bold=True, size=10, numfmt=MULT, align=S.center)
write(comps, f'G{exit_ff_row}', f"=(E{CM['ebitda30']}*E{exit_ff_row}+E{CM['cash']})/E{CM['sh']}", S.BLACK, bold=True, size=10, numfmt=MONEY, align=S.center)
write_assumption_docs(comps, exit_ff_row, DJ, DS, DC, 'dcf_exitm', D.JUST, D.ASSUMPTION_SRC,
                      internal_location=f"'DCF'!E{VR['exitm']}", hints=D.SOURCE_HINT)
rr[0] += 1
ff_pe("P / E (FY2026E EPS)", 10.0, 18.0)
# DCF range row references DCF sheet outputs
write(comps, f'A{rr[0]}', "DCF (bear \u2013 bull)", S.BLACK, size=10, align=S.left_indent)
write(comps, f'E{rr[0]}', "\u2014", S.BLACK, size=10, align=S.center)
write(comps, f'F{rr[0]}', "\u2014", S.BLACK, size=10, align=S.center)
write(comps, f'G{rr[0]}', f"=Scenarios!F{pt_row}", S.BLACK, size=10, numfmt=MONEY, align=S.center)
write(comps, f'H{rr[0]}', f"=Scenarios!H{pt_row}", S.BLACK, size=10, numfmt=MONEY, align=S.center)
rr[0] += 2
write(comps, f'A{rr[0]}', "Current price", S.BLACK, bold=True, size=11, align=S.left_indent)
write_reported(comps, f'E{rr[0]}', D.MKT['price'], D.SOURCES["nasdaq_quote"], bold=True, size=11, numfmt=MONEY)
write_source_with_ctrl_f(comps, f'{DS}{rr[0]}', f'{DC}{rr[0]}', "NASDAQ", D.SOURCES["nasdaq_quote"],
                         D.REPORTED_HINTS["nasdaq"])
rr[0] += 1
write(comps, f'A{rr[0]}', "Note: peer set includes NKE, DECK, ONON, adidas, VFC; multiples are analyst ranges (red).",
      S.BLACK, italic=True, size=8, align=S.left_indent)

# Link DCF exit multiple to Comps peer build + peer table
write_internal_link(dcf, f'M{VR["exitm"]}', 'Comps: Exit Multiple Build', f"'Comps'!A{EM['selected']}")

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("Saved", os.path.abspath(OUT))
